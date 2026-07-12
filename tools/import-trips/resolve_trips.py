import argparse
import openpyxl
import re
import json
import difflib
import uuid
from datetime import datetime, date
from collections import Counter

parser = argparse.ArgumentParser(
    description="Resolve Truck Loading Info.xlsx rows into importable trip/discharge records "
                "against a target environment's reference data (Trucks/Drivers/Stations)."
)
parser.add_argument("--xlsx", default="/Users/mustaphaaliyu/Projects/mcgmis/Shared/Truck Loading Info.xlsx",
                     help="Path to the source Excel workbook.")
parser.add_argument("--ref-dir", default="/tmp",
                     help="Directory containing db_stations.txt, db_trucks.txt, db_drivers.txt "
                          "for the TARGET environment (local or prod). Generate these with "
                          "prod-runbook/02_export_reference_data.sh pointed at that environment "
                          "before running this against prod.")
parser.add_argument("--out", default="import_data.json",
                     help="Output path for the resolved import JSON.")
args = parser.parse_args()

XLSX = args.xlsx
MONTHS = ['February 2025', 'March 2025', 'April 2025', 'May 2025', 'June 2025', 'July 2025', 'August 2025', 'September 2025']

# ---------- Load DB reference data ----------
def load_pipe_file(path):
    # Format: header line, "---" separator line, N data lines, "(N rows)" footer.
    rows = []
    with open(path) as f:
        lines = [l.rstrip("\n") for l in f]
    for line in lines[2:-1]:
        rows.append([c.strip() for c in line.split("|")])
    return rows

db_stations = load_pipe_file(f"{args.ref_dir}/db_stations.txt")
db_trucks = load_pipe_file(f"{args.ref_dir}/db_trucks.txt")
db_drivers = load_pipe_file(f"{args.ref_dir}/db_drivers.txt")

truck_by_plate = {}
for r in db_trucks:
    if len(r) >= 2:
        key = re.sub(r"\s+", "", r[1].upper())
        truck_by_plate[key] = r[0]

station_by_name_upper = {r[1].strip().upper(): r[0] for r in db_stations if len(r) >= 2 and r[1].strip()}

driver_records = [(r[0], r[1].strip(), r[2].strip()) for r in db_drivers if len(r) >= 3]

# NOTE: Jul-Sep 2025 trips that previously existed in the DB have been deleted
# (per user instruction: wipe Jul/Aug/Sep and replace wholesale with the Excel
# data instead of trying to dedup/merge). No dedup step needed anymore.

# ---------- Manual mapping tables (from analysis + user confirmation) ----------
LOADING_POINT_MAP = {
    "DANGOTE REFINERY": "Dangote Refinery",
    "NIPCO IBAFO": "NIPCO Ibafo",
    "APAPA TERMINAL 1": "NRL Apapa Terminal 1",
    "NIPCO APAPA": "NIPCO Apapa Terminal",
    "APTERM2 - JV": "Apapa Terminal 2 JV",
    "OLOGBO": "Ologbo LPG",
    "ILASAMAJA": "Ilasamanja CNG Mother Station",
    "AJAOKUTA": "NIPCO Ajaokuta CNG Mother Station",
}

# High-confidence set + user-confirmed set (all keyed by UPPERCASE normalized discharge location)
DISCHARGE_LOCATION_MAP = {
    "IKEJA AVIATION": "Ikeja Aviation",
    "ABUJA AVIATION": "Abuja Aviation",
    "NNPC GUZAPE": "NNPC Guzape",
    "1ST AVENUE GWARINPA": "NNPC 1St Avenue Gwarinpa",
    "NNPC 1ST AVENUE GWARINPA": "NNPC 1St Avenue Gwarinpa",
    "NNPC 1ST AVE GWARINPA": "NNPC 1St Avenue Gwarinpa",
    "NNPC OBAFEMI AWOLOWO JABI": "NNPC Obafemi Awolowo Way Jabi",
    "NNPC OBAFEMI AWOLOWO WAY JABI": "NNPC Obafemi Awolowo Way Jabi",
    "NNPC KUGBO": "NNPC Kugbo",
    "NNPC DEI-DEI": "NNPC Dei-Dei",
    "DEI-DEI": "NNPC Dei-Dei",
    "DEI DEI": "NNPC Dei-Dei",
    "NNPC DIKWA ROAD KUBWA": "NNPC Dikwa Road Kubwa",
    "PRESIDENTIAL VILLA": "Presidential Villa C-In-C",
    "GADUWA": "NNPC Gaduwa",
    "GDUWA": "NNPC Gaduwa",
    "USHAFA": "NNPC Station Ushafa",
    "KADUNA": "Olam, Kaduna",
    "APAPA": "HOGL-APAPA",
    "BENIN": "NBC Benin",
    "ENUGU": "NBC Enugu",
    "MASAKA": "Masaka LPG Plant",
    "MABUSHI": "NNPC Mabushi",
    "WARRI": "Warri LPG Plant",
    # Resolved this round:
    "SECTOR CENTRE": "NNPC Princess & Princess Junction (Sector Center)",
    "SECTOR CENTER": "NNPC Princess & Princess Junction (Sector Center)",
    "ABUJA SEC CENTRE": "NNPC Princess & Princess Junction (Sector Center)",
    "O.O WAY": "NNPC Station Olusegun Obasanjo Way, Wuse Zone 1",
    "O.O. WAY": "NNPC Station Olusegun Obasanjo Way, Wuse Zone 1",
    "OO": "NNPC Station Olusegun Obasanjo Way, Wuse Zone 1",
    "OLUSEGUN OBJ WAY": "NNPC Station Olusegun Obasanjo Way, Wuse Zone 1",
    "GWAGWALADA": "Adenle Abdulkadir",
    "IKEJA AIRPORT": "Ikeja Aviation",
    "C.AVE": "NNPC Constitution Avenue",
    "C. AVE": "NNPC Constitution Avenue",
    "CONST.AVE": "NNPC Constitution Avenue",
    "CONST. AVE": "NNPC Constitution Avenue",
    "CONST AVE": "NNPC Constitution Avenue",
    "CONST .AVE": "NNPC Constitution Avenue",
    "CONSTITUTION AVE": "NNPC Constitution Avenue",
    "CONSTITUTIONAL AVE.": "NNPC Constitution Avenue",
    "NOI JABI": "NNPC Ngozi Okonjo-Iweala Way Utako",
    "NNPC NOI JABI": "NNPC Ngozi Okonjo-Iweala Way Utako",
    "NNPC NOI UTAKO": "NNPC Ngozi Okonjo-Iweala Way Utako",
    "HERBERT MACAULAY Z3": "NNPC Station Wuse Zone 3 (Ahmed Salihi)",
    "HERBERT MACAULAY Z6": "NNPC Station Wuse Zone 6",
    "LIFE CAMP": "NNPC Life-Camp",
    "LIFE-CAMP": "NNPC Life-Camp",
    "KUBWA": "NNPC Dikwa Road Kubwa",
    "USHAFA BWARI": "NNPC USHAFA-BWAWRI RD S/S (ABUBAKAR MUHAMMAD)",
    "TASHA BWARI": "NNPC World Energy Tasha-Bwari Way",
    "SHOPRITE PYAKASA": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "NYANYA": "PATKEM ADMIRALTY (NNPC New Nyanya)",
    # "RIVERS" intentionally omitted -> falls to no-match -> discharge skipped per user instruction
    # --- User-confirmed mappings from station_mapping_review.csv round ---
    "ABUJA AIRPORT": "Abuja Aviation",
    "CCECC NIG LTD": "CCECC LAGOS",
    "NNPC LUGBE": "NNPC Carwash Bus Stop Lugbe (BAYSCOM ENERGY)",
    "NNPC DURUMI": "NNPC Durumi AIS Axis",
    "NNPC HERBERT MACAULAY WUSE Z6": "NNPC Station Wuse Zone 6",
    "NNPC MEGA STATION ABUJA": "NNPC Mega Station Zone 1",
    "FOLA": "NNPC CNG Fola Agoro Station",
    "JOS": "Jos LPG Plant",
    "KATAMPE - JAHI": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC ALAPERE": "NNPC Alapere Station (Funmilayo Alebioso)",
    "NNPC GARKI 2": "NNPC Ahmadu Bello Way Garki 2",
    "NNPC KATAMPE-JAHI AXIS MM WAY": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC MEGA STATION OO": "NNPC Station Olusegun Obasanjo Way, Wuse Zone 1",
    "NNPC WORLD ENERGY TASHA BWARI": "NNPC World Energy Tasha-Bwari Way",
    "PRINCE & PRINCESS": "NNPC Sector Center, Kaura (Prince & Princess)",
    "AGEGE": "NNPC Station CNG Station Agege",
    "C/AVE": "NNPC Constitution Avenue",
    "GARKI 2": "NNPC Ahmadu Bello Way Garki 2",
    "GARKI AREA 11": "NNPC Ahmadu Bello Way Garki 2",
    "IBADAN": "NNPC CNG Queen Elizabeth",
    "KATAMPE EXT": "NNPC Katampe Extension (Aminu Jega)",
    "LYUSCHI": "Idu LPG Plant Lyushi",
    "LIFE CAMP - JABI": "NNPC Obafemi Awolowo Way Jabi",
    "MONACO PETROLEUM MARARRABA-KARU EXP WAY": "NNPC Mararaba Near Old Karu Rd (Monaco Oil)",
    "NASARAWA": "PATKEM ADMIRALTY (NNPC New Nyanya)",
    "NBC PH": "NBC Warri Depot",
    "NNPC AIRPORT ROAD ADJACENT SHOPRITE (SEDDI KEMIL)": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "NNPC GARKI 2 (HADIZAT JIBRIN)": "NNPC Ahmadu Bello Way Garki 2",
    "NNPC KATAMPE-JAHI JUNCTION": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC LUGBE (BAYSCOM)": "NNPC Carwash Bus Stop Lugbe (BAYSCOM ENERGY)",
    "NNPC LUGBE AFTER DUNAMIS": "NNPC Airport Road Lugbe After Dunamis Church",
    "NNPC MEGA": "NNPC Mega Station Zone 1",
    "NNPC MM WAY KATAMPE-JAHI AXIS": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC MARARABA (MONACO OIL), NEAR OLD KARU WAY": "NNPC Mararaba Near Old Karu Rd (Monaco Oil)",
    "NNPC MARARABA OLD KARU AXIS": "NNPC Mararaba Near Old Karu Rd (Monaco Oil)",
    "NNPC MARINA": "NNPC Station Marina (Orji Tinubu)",
    "NNPC MEGA STATION WUSE Z1": "NNPC Mega Station Zone 1",
    "NNPC MONACO KARU": "NNPC Mararaba Near Old Karu Rd (Monaco Oil)",
    "NNPC NYANYA-KARSHI ROAD": "NNPC Jikwoyi Nyanya-Karshi Road",
    "NNPC USHAFA BWARI": "NNPC USHAFA-BWAWRI RD S/S (ABUBAKAR MUHAMMAD)",
    "NNPC WUSE ZONE 3 (AHMED SALIHI)": "NNPC Station Wuse Zone 3 (Ahmed Salihi)",
    "NNPC KUGBO KEFFI ROAD": "NNPC Kugbo",
    "OA JABI": "NNPC Obafemi Awolowo Way Jabi",
    "OO MEGA STATION": "NNPC Mega Station Zone 1",
    "OTTA": "NNPC CNG Ota Station",
    "PHC": "Port Harcourt Aviation",
    "PRINCESS & PRINCESS": "NNPC Princess & Princess Junction (Sector Center)",
    "SOKOTO": "Sokoto LPG Plant",
    "SHOPRITE": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "ABUJA": "Abuja Aviation",
    "AK AVE LIFE-CAMP": "NNPC Life-Camp",
    "ASCO JUNCTION KOGI": "NNPC ASCO Junction (O Raji)",
    "APAPA LPG PLANT": "NRL Terminal 1 LPG Plant",
    "APAPA T1 LPG PLANT": "NRL Terminal 1 LPG Plant",
    "BAKKA OIL ABUJA-LOKOJA EXP WAY, KWALI, ABUJA": "Bakka Oil",
    "CA CONST AVE": "NNPC Constitution Avenue",
    "CENTRAL AREA CONST AVE": "NNPC Constitution Avenue",
    "CONST/AVE": "NNPC Constitution Avenue",
    "CROWN FLOUR MILL WARRI": "Crown Flour Mill Ltd Warri",
    "GUZAPE": "NNPC Guzape",
    "IDU ABUJA": "Idu LPG Plant Lyushi",
    "JABI - LIFE CAMP": "NNPC Obafemi Awolowo Way Jabi",
    "KATAMPE-JAHI": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "KUBWA DIKWA ROAD": "NNPC Dikwa Road Kubwa",
    "LYUSHY": "Idu LPG Plant Lyushi",
    "LIFE CAMP - JABI BRIDGE": "NNPC Obafemi Awolowo Way Jabi",
    "MARINA": "NNPC Station Marina (Orji Tinubu)",
    "MMW GWARINPA/JAHI": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "MEGA STATION": "NNPC Mega Station Zone 1",
    "MEGA STATION JAHI YO WAY": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NIGERIA PETROLEUM DEV COMPANY": "Nigerian Petroleum Development Company",
    "NNPC ASCO JUNCTION, OHUMEME-LOKOJA ROAD, KOGI": "NNPC ASCO Junction (O Raji)",
    "NNPC ASCO JUNCTION, AJAOKUTA (O RAJI)": "NNPC ASCO Junction (O Raji)",
    "NNPC ABKR KOKO AVE LIFE-CAMP": "NNPC Life-Camp",
    "NNPC AIRPORT ROAD ADJ SHOPRITE (SADDI KEMIL)": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "NNPC AIRPORT ROAD SHOPRITE (SEDDI KEMIL)": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "NNPC AJAOKUTA (ALL DE NATIONS), AJAOKUTA-ENUGU ROAD": "ALL DE NATION OIL & GAS",
    "NNPC AJAOKUTA-AYANGBA ROAD, KOGI": "NNPC Ajaokuta-Ayangba Way Kogi (Mohammed Anyagba)",
    "NNPC AJAOKUTA-AYANGBA WAY KOGI": "NNPC Ajaokuta-Ayangba Way Kogi (Mohammed Anyagba)",
    "NNPC ALLOMA OFU KOGI": "NNPC Allomo, Kogi",
    "NNPC ALLOMA, KOGI": "NNPC Allomo, Kogi",
    "NNPC ALONG ABJ KEFFI ROAD NYANYA": "PATKEM ADMIRALTY (NNPC New Nyanya)",
    "NNPC ANYIGBA KOGI": "NNPC Ajaokuta-Ayangba Way Kogi (Mohammed Anyagba)",
    "NNPC AYANGBA (MOHAMMED AYANGBA)": "NNPC Ajaokuta-Ayangba Way Kogi (Mohammed Anyagba)",
    "NNPC AYANGBA KOGI": "NNPC Ajaokuta-Ayangba Way Kogi (Mohammed Anyagba)",
    "NNPC BALEWA ROAD, ANKA, KOGI": "NNPC Balewa Road, Ankpa, Kogi (Alfa Allied)",
    "NNPC BAYSCOM ENERGY LTD LUGBE": "NNPC Carwash Bus Stop Lugbe (BAYSCOM ENERGY)",
    "NNPC BWARI": "NNPC World Energy Tasha-Bwari Way",
    "NNPC CHECKPOINT, OKENE-AUCHI RD, OKENE": "NNPC Checkpoint Okene",
    "NNPC CONST AVENUE CBD": "NNPC Constitution Avenue",
    "NNPC DEI DEI": "NNPC Dei-Dei",
    "NNPC DEI-DEI SS (IDRIS ZANGI)": "NNPC Dei-Dei",
    "NNPC DEIDEI": "NNPC Dei-Dei",
    "NNPC DIKWA RD KUBWA (SUNDAY DADA)": "NNPC Dikwa Road Kubwa",
    "NNPC DIKWA ROAD, KUBWA": "NNPC Dikwa Road Kubwa",
    "NNPC EJULE/ALLOMA - ENUGU ROAD, KOGI": "NNPC Allomo, Kogi",
    "NNPC FELELE, LOKOJA-ABJ EXP WAY, KOGI": "NNPC Felele Service Station (Ahmed Musa)",
    "NNPC GANNAJA ROAD, LOKOJA, KOGI": "NNPC Gannaja Road Lokoja-Ajaokuta Way",
    "NNPC GUDU JUNCTION": "NNPC Abdussalami Abubakar Way Gudu",
    "NNPC HERBERT M Z6 (B TANKO)": "NNPC Station Wuse Zone 6",
    "NNPC HERBERT MACAULAY JABI": "NNPC Obafemi Awolowo Way Jabi",
    "NNPC HERBERT MACAULAY WUSE Z1": "NNPC Station Wuse Zone 3 (Ahmed Salihi)",
    "NNPC HERBERT MACAULAY WUSE Z3": "NNPC Station Wuse Zone 3 (Ahmed Salihi)",
    "NNPC HERBERT MACAULAY Z3": "NNPC Station Wuse Zone 3 (Ahmed Salihi)",
    "NNPC HERBERT MACAULAY Z6": "NNPC Station Wuse Zone 6",
    "NNPC HERBERT MACAULAY Z6 (B TANKO)": "NNPC Station Wuse Zone 6",
    "NNPC HERBERT MACAULAY ZONE # ABUJA": "NNPC Station Wuse Zone 3 (Ahmed Salihi)",
    "NNPC IDAH AYANGBA ROAD, KOGI": "NNPC Ajaokuta-Ayangba Way Kogi (Mohammed Anyagba)",
    "NNPC IYARA/IJUMU, KOGI": "NNPC Iyara, Ijumu, Kogi",
    "NNPC JAHI KATAMPE BRIDGE": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC JIKWOYI": "NNPC Jikwoyi Nyanya-Karshi Road",
    "NNPC JIKWOYI, KARSHI-NYANYA WAY": "NNPC Jikwoyi Nyanya-Karshi Road",
    "NNPC KOGI(": "NNPC Mega Station Lokoja",
    "NNPC KUBGO": "NNPC Kugbo",
    "NNPC KATAMPE-JAHI AXIS": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC KATAMPE-JAHI AXIS (YMI PETROLEUM)": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC KATAMPE-JAHI AXIS, MM WAY": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC KUBWA (SARCO OIL)": "NNPC Kubwa Service Station (Sarco Pet & Gas)",
    "NNPC KUBWA (SARCO PET & GAS)": "NNPC Kubwa Service Station (Sarco Pet & Gas)",
    "NNPC LIFE-CAMP ABK KOKO AVE (IBRAHIM BATENGI)": "NNPC Life-Camp",
    "NNPC LIFECAMP ABKR KOKO AVE": "NNPC Life-Camp",
    "NNPC LOKOJA-OKENE ROAD": "Haba Engineering Services Ltd",
    "NNPC LOKOJA/OKENE RD": "Haba Engineering Services Ltd",
    "NNPC LOKOJA/OKENE WAY FELELE (HABA ENGR)": "Haba Engineering Services Ltd",
    "NNPC LOKONGOMA PHASE II LOKOJA (HAMMAJOY)": "NNPC Lokongoma Phase II Service Station Lokoja (Hammajoy)",
    "NNPC LUGBE ATV SIDE": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "NNPC LUGBE BEFORE DUNAMIS": "NNPC Carwash Bus Stop Lugbe (BAYSCOM ENERGY)",
    "NNPC LUGBE CAR WASH": "NNPC Carwash Bus Stop Lugbe (BAYSCOM ENERGY)",
    "NNPC MABUSHI BY VIO": "NNPC Mabushi",
    "NNPC MM WAY JAHI-KATAMPE AXIS": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC MM WAY JAHI-KATAMPE AXIS (YMI)": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC MM WAY KATAMPE-JAHI AXIS (YMI)": "NNPC Katampe-Jahi Axis, Kubwa Express Way (YMI Petroleum Ltd)",
    "NNPC MM WAY LOKOJA": "NNPC MM Way Service Station (Umar Yakubu)",
    "NNPC MARARABA CLOSE TO OLD KARU ROAD": "NNPC Mararaba Near Old Karu Rd (Monaco Oil)",
    "NNPC MARARABA NEAR OLD KARU RD, ABJ-KEFFI RD": "NNPC Mararaba Near Old Karu Rd (Monaco Oil)",
    "NNPC MARARABA-KARU AXIS": "NNPC Mararaba Near Old Karu Rd (Monaco Oil)",
    "NNPC MARARARBA NEAR OLD KARU (MONACO OIL)": "NNPC Mararaba Near Old Karu Rd (Monaco Oil)",
    "NNPC MEGA LOKOJA": "NNPC Mega Station Lokoja",
    "NNPC MEGA STATION ABJ-LOKOJA WAY KWALI": "Bakka Oil",
    "NNPC MEGA STATION ADO EKITI": "NNPC Mega Ado-Ekiti",
    "NNPC MEGA STATION AJAOKUTA": "NNPC Mega Station Lokoja",
    "NNPC MEGA STATION LOKOJA-OKENE ROAD, KOGI": "NNPC Mega Station Lokoja",
    "NNPC MEGA STATION OO ABUJA": "NNPC Mega Station Zone 1",
    "NNPC MEGA STATION OO WUSE Z1": "NNPC Mega Station Zone 1",
    "NNPC MEGA STATION WUSE Z1 ABUJA": "NNPC Mega Station Zone 1",
    "NNPC NOI": "NNPC Ngozi Okonjo-Iweala Way Utako",
    "NNPC NOI UTAKO S/S (GAFAR WORLDWIDE)": "NNPC Ngozi Okonjo-Iweala Way Utako",
    "NNPC OFU NEW LAYOUT, KOGI": "NNPC Ofu New Layout Service Station (Danjuma Akowe)",
    "NNPC OBAFEMI AWOLOW JABI": "NNPC Obafemi Awolowo Way Jabi",
    "NNPC OBAFEMI AWOLOWO LIFE-CAMP BRIDGE JABI": "NNPC Obafemi Awolowo Way Jabi",
    "NNPC OHUMEME-LOKOJA ROAD, OKENE": "NNPC ASCO Junction (O Raji)",
    "NNPC OKENE/AJAKUTA ROAD (SIYAKA SANNI)": "NNPC Ajaokuta/Okene Road Service Station (Siyaka Sanni)",
    "NNPC OPPOSITE SHOPRITE LUGBE": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "NNPC PRINCE & PRINCESS (SECTOR CENTER)": "NNPC Princess & Princess Junction (Sector Center)",
    "NNPC PYAKASA": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "NNPC SEDDI KEMIL SHOPRITE AIRPORT ROAD (3SK PROPERTIES)": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "NNPC STATION FELELE, KOGI": "NNPC Felele Service Station (Ahmed Musa)",
    "NNPC STATION KABUSA": "NNPC Kabusa",
    "NNPC TASHA BWARI RD (ABUBAKAR MUHAMMAD)": "NNPC USHAFA-BWAWRI RD S/S (ABUBAKAR MUHAMMAD)",
    "NNPC USHAFA": "NNPC Station Ushafa",
    "NNPC USHAFA (BORMI RESOURCES)": "NNPC USHAFA-BWAWRI RD S/S (ABUBAKAR MUHAMMAD)",
    "NNPC USHAFA - BWARI ROAD": "NNPC USHAFA-BWAWRI RD S/S (ABUBAKAR MUHAMMAD)",
    "NNPC USHAFA BWARI ROAD": "NNPC USHAFA-BWAWRI RD S/S (ABUBAKAR MUHAMMAD)",
    "NNPC USHAFA TASHA BWARI": "NNPC World Energy Tasha-Bwari Way",
    "NNPC WORD ENERGY TASHA BWARI": "NNPC World Energy Tasha-Bwari Way",
    "NNPC WORLD ENERGY BWARI": "NNPC World Energy Tasha-Bwari Way",
    "NNPC WUSE ZONE 6 (B TANKO)": "NNPC Station Wuse Zone 6",
    "NEW NYANYA": "PATKEM ADMIRALTY (NNPC New Nyanya)",
    "NYANYA - KARSHI ROAD": "NNPC Jikwoyi Nyanya-Karshi Road",
    "NYANYA-KARSHI": "NNPC Jikwoyi Nyanya-Karshi Road",
    "O.O.WAY": "NNPC Station Olusegun Obasanjo Way, Wuse Zone 1",
    "OLUSEGUN": "NNPC Station Olusegun Obasanjo Way, Wuse Zone 1",
    "OPP SHOPRITE": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "PH AVIATION": "Port Harcourt Aviation",
    "PORTHACOURT": "Port Harcourt Aviation",
    "PORTHARCOURT": "Port Harcourt Aviation",
    "PASERE ABJ-LKJ ROAD": "Adenle Abdulkadir",
    "PRINCE & PRINCES": "NNPC Sector Center, Kaura (Prince & Princess)",
    "SHOPRITE LUGBE": "NNPC Airport Road Adjascent Shoprite (Seddi Kemil)",
    "US EMBASSY": "US Embassy Abuja",
    "ZUBA": "NNPC Zuba",
    "PH": "Port Harcourt Aviation",
}

DISCARD_STATIONS = {
    "RD SULEJA",
    "NNPC STATION",
}

# Addresses for genuinely-new stations that the user confirmed manually
# (keyed by normalize_key of the display name used in new_stations output).
KNOWN_STATION_ADDRESSES = {
    "GOODONE CARTON NIG LTD": {"location": "Sango-Ota", "state": "Ogun", "contact_address": None},
    "DAHUA PAPER COMPANY": {"location": "Sango-Ota", "state": "Ogun", "contact_address": None},
    "DSS": {"location": "AMAC", "state": "Abuja", "contact_address": None},
}

def normalize_key(s):
    return re.sub(r"\s+", " ", str(s).strip()).upper()

def _direct_lookup(key, extra_map):
    if extra_map and key in extra_map:
        target_name = extra_map[key]
        sid = station_by_name_upper.get(target_name.upper())
        if sid:
            return sid
    return station_by_name_upper.get(key)

def resolve_station(name, extra_map=None):
    """Returns (station_id_or_None, resolved_name_for_new_station_or_None).

    Tries the raw text first, then a small set of safe textual normalizations
    (strip trailing parenthetical agent-name suffix, hyphen/space normalize,
    first-segment of a "/"-or-","-joined compound) against the same map/DB —
    never against fuzzy/guessed names — before giving up and flagging as new.
    """
    if not name:
        return None, None
    raw = str(name).strip()
    key = normalize_key(raw)
    if key == "RIVERS" or key in DISCARD_STATIONS:
        return None, None  # explicit skip, no auto-create

    sid = _direct_lookup(key, extra_map)
    if sid:
        return sid, None

    candidates = []
    no_paren = re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()
    if no_paren and no_paren != raw:
        candidates.append(no_paren)
    for c in list(candidates):
        alt = re.sub(r"\s+", " ", c.replace("-", " ")).strip()
        if alt and alt not in candidates:
            candidates.append(alt)
    for sep in ("/", ","):
        if sep in raw:
            first = raw.split(sep)[0].strip()
            if first and first not in candidates:
                candidates.append(first)

    for cand in candidates:
        ckey = normalize_key(cand)
        if ckey == "RIVERS":
            continue
        sid = _direct_lookup(ckey, extra_map)
        if sid:
            return sid, None

    # Not matched anywhere -> auto-create using cleaned raw text
    return None, raw

def match_truck(plate):
    if not plate:
        return None
    key = re.sub(r"\s+", "", str(plate).upper())
    return truck_by_plate.get(key)

def match_driver(name_str):
    if not name_str:
        return None
    s = re.sub(r"^\s*\d+\s*-\s*", "", str(name_str)).strip().upper()
    toks = s.split()
    if not toks:
        return None
    best = None
    best_score = 0.0
    for did, fn, ln in driver_records:
        db_toks = f"{fn} {ln}".upper().split()
        score = 0.0
        for t in toks:
            best_tok = max((difflib.SequenceMatcher(None, t, dt).ratio() for dt in db_toks), default=0)
            score += best_tok
        avg = score / len(toks)
        if avg > best_score:
            best_score = avg
            best = did
    if best_score >= 0.90:  # confirmed spelling-variant threshold
        return best
    return None  # blank rather than risk a wrong person

DATE_RE = re.compile(r"\d{1,2}/\d{1,2}/\d{2,4}")

def parse_dates_from_field(value, fallback_count):
    """Return a list of date() objects, length == fallback_count (reusing single date if needed)."""
    if value is None:
        return [None] * fallback_count
    if isinstance(value, datetime):
        return [value.date()] * fallback_count
    s = str(value)
    matches = DATE_RE.findall(s)
    dates = []
    for m in matches:
        parts = re.split(r"/", m)
        if len(parts) != 3:
            continue
        mo, da, yr = parts
        try:
            mo, da = int(mo), int(da)
            yr = int(yr)
            if yr < 100:
                yr += 2000
            dates.append(date(yr, mo, da))
        except ValueError:
            continue
    if not dates:
        return [None] * fallback_count
    if len(dates) == fallback_count:
        return dates
    if len(dates) == 1:
        return dates * fallback_count
    # ambiguous count -> reuse first date for all rather than guess pairing
    return [dates[0]] * fallback_count

def to_float(s):
    try:
        return float(str(s).strip().replace(",", ""))
    except (ValueError, TypeError):
        return None

def parse_discharges(loc_raw, qty_raw, date_raw):
    """
    Returns (discharges: list of {location_raw, quantity, date}, parse_ok: bool)
    parse_ok False means: could not confidently split -> caller should skip discharge
    creation but keep raw text for notes.
    """
    if qty_raw is None:
        return [], True  # no discharge data at all; not an error, just nothing to record
    qty_str = str(qty_raw)
    qty_parts_raw = [p.strip() for p in qty_str.split('/') if p.strip()]
    qty_vals = [to_float(p) for p in qty_parts_raw]
    if not qty_parts_raw or any(v is None for v in qty_vals):
        return [], False  # unparseable quantity token(s)

    n_qty = len(qty_vals)
    loc_str = str(loc_raw).strip() if loc_raw else ""
    locs_split = [p.strip() for p in loc_str.split('/') if p.strip()]

    if n_qty == 1:
        locations = [loc_str] if loc_str else [""]
    elif len(locs_split) == n_qty:
        locations = locs_split
    elif len(locs_split) == 1:
        locations = locs_split * n_qty
    elif len(locs_split) == 0:
        return [], False  # quantities given but no location at all
    else:
        return [], False  # genuine count mismatch, can't safely pair

    dates = parse_dates_from_field(date_raw, n_qty)
    discharges = []
    for loc, qty, dt in zip(locations, qty_vals, dates):
        discharges.append({"location_raw": loc, "quantity": qty, "date": dt.isoformat() if dt else None})
    return discharges, True

def parse_single_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    m = DATE_RE.search(str(value))
    if not m:
        return None
    parts = m.group(0).split("/")
    if len(parts) != 3:
        return None
    mo, da, yr = parts
    try:
        mo, da, yr = int(mo), int(da), int(yr)
        if yr < 100:
            yr += 2000
        return date(yr, mo, da).isoformat()
    except ValueError:
        return None

def determine_product_enum(product_raw, loading_point_raw):
    p = (product_raw or "").strip().upper()
    lp = normalize_key(loading_point_raw) if loading_point_raw else ""
    if p == "CNG":
        if lp == "AJAOKUTA":
            return "CngAbuja"
        elif lp in ("ILASAMAJA", "NIPCO IBAFO"):
            return "CngLagos"
        else:
            return "CngAbuja"  # default fallback for the 1 anomalous Dangote-Refinery CNG row
    return {"PMS": "PMS", "ATK": "ATK", "LPG": "LPG", "AGO": "AGO", "PAO": "PAO"}.get(p, p)

# ---------- Main extraction ----------
wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)

results = []
skipped_no_truck = 0
new_stations_needed = {}  # upper-key -> display name

for month in MONTHS:
    ws = wb[month]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header = [h.strip() if isinstance(h, str) else h for h in header]
    col_idx = {}
    for i, h in enumerate(header):
        if h and h not in col_idx:
            col_idx[h] = i

    def get(row, name):
        idx = col_idx.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if get(row, 'Loading Date') is None and get(row, 'Truck Number') is None:
            continue

        truck_plate = get(row, 'Truck Number')
        truck_id = match_truck(truck_plate)
        if not truck_id:
            skipped_no_truck += 1
            continue

        loading_date_raw = get(row, 'Loading Date')
        loading_date = parse_single_date(loading_date_raw)
        if not loading_date:
            continue  # can't create a Trip without a date

        waybill = get(row, 'Waybill Number')
        waybill_str = str(waybill).strip() if waybill is not None else ""

        product_raw = get(row, 'Product')
        loading_point_raw = get(row, 'Loading Point')
        product_enum = determine_product_enum(product_raw, loading_point_raw)

        loading_point_id, new_lp_name = resolve_station(loading_point_raw, LOADING_POINT_MAP)
        if new_lp_name:
            new_stations_needed[normalize_key(new_lp_name)] = new_lp_name

        driver_name_raw = get(row, 'Driver Name')
        driver_id = match_driver(driver_name_raw)

        dispatch_qty = to_float(get(row, 'Dispatch Quatity (MT/SCM/KG/LTR)'))

        discharge_location_raw = get(row, 'Discharge Location')
        discharge_qty_raw = get(row, 'Discharged Quatity (MT/SCM/KG/LTR)')
        discharge_date_raw = get(row, 'Discharge Date')
        discharges_parsed, parse_ok = parse_discharges(discharge_location_raw, discharge_qty_raw, discharge_date_raw)

        resolved_discharges = []
        discharge_note = None
        if not parse_ok:
            discharge_note = f"UNPARSED DISCHARGE DATA - loc={discharge_location_raw!r} qty={discharge_qty_raw!r} date={discharge_date_raw!r}"
        else:
            for d in discharges_parsed:
                sid, new_name = resolve_station(d["location_raw"], DISCHARGE_LOCATION_MAP)
                if new_name:
                    new_stations_needed[normalize_key(new_name)] = new_name
                if sid is None and new_name is None:
                    # e.g. "Rivers" or blank -> skip this specific discharge, note it
                    discharge_note = (discharge_note or "") + f"; SKIPPED DISCHARGE (no station): {d['location_raw']!r} qty={d['quantity']}"
                    continue
                resolved_discharges.append({
                    "station_id": sid,
                    "new_station_name": new_name,
                    "quantity": d["quantity"],
                    "date": d["date"],
                })

        results.append({
            "month": month,
            "truck_id": truck_id,
            "driver_id": driver_id,
            "loading_date": loading_date,
            "waybill_no": waybill_str or None,
            "dispatch_qty": dispatch_qty,
            "product": product_enum,
            "status": "Closed",
            "loading_depot_id": loading_point_id,
            "destination": get(row, 'Destination'),
            "elock_status": get(row, 'E-lock Status'),
            "arrived_atv": get(row, 'Arrived ATV?'),
            "atv_arrival_date": parse_single_date(get(row, 'ATV Arrival Date')),
            "arrived_station": get(row, 'Arived Station?'),
            "station_arrival_date": parse_single_date(get(row, 'Station Arrival Date')),
            "return_date": parse_single_date(get(row, 'Return Date')),
            "remarks": get(row, 'Remarks'),
            "discharge_note": discharge_note,
            "discharges": resolved_discharges,
        })

print(f"Rows resolved for import: {len(results)}")
print(f"Skipped (truck not found): {skipped_no_truck}")
print(f"New stations needed (auto-create): {len(new_stations_needed)}")

out = {
    "trips": results,
    "new_stations": [
        {"name": name, "address": KNOWN_STATION_ADDRESSES.get(normalize_key(name))}
        for name in sorted(new_stations_needed.values())
    ],
}
with open(args.out, "w") as f:
    json.dump(out, f, indent=2, default=str)

print(f"\nWritten to {args.out}")

# quick sanity stats
no_discharge_at_all = sum(1 for r in results if not r["discharges"] and not r["discharge_note"])
with_note = sum(1 for r in results if r["discharge_note"])
print(f"\nTrips with zero discharges (no data ever provided): {no_discharge_at_all}")
print(f"Trips with a discharge_note (partial issue flagged): {with_note}")
